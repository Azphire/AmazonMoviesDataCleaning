from openpyxl import Workbook
from openpyxl import load_workbook

movies = load_workbook(
    'D:\\Projects\\PythonProjects\\DataCleaning\\data\\FinalMovies.xlsx')  # 读取xlsx中第一个sheet
sheet = movies['Sheet']
rows = sheet.max_row
columns = sheet.max_column

# 存储输出文件：导演
wb = Workbook()
ws = wb.active
ws.append(['id', '导演'])

for i in range(2, rows + 1):
    try:
        directors = sheet.cell(row=i, column=5).value.split(',')
        for director in directors:
            myRow = [sheet.cell(row=i, column=1).value, director.strip()]
            ws.append(myRow)
    except:
        pass

wb.save('D:\\Projects\\PythonProjects\\DataCleaning\\data\\Directors.xlsx')