from openpyxl import Workbook
from openpyxl import load_workbook

movies = load_workbook(
    'D:\\Projects\\PythonProjects\\DataCleaning\\data\\FinalMovies.xlsx')  # 读取xlsx中第一个sheet
sheet = movies['Sheet']
rows = sheet.max_row
columns = sheet.max_column

# 存储输出文件：演员
wb = Workbook()
ws = wb.active
ws.append(['id', '演员'])

for i in range(2, rows + 1):
    try:
        actors = sheet.cell(row=i, column=6).value.split(',')
        for actor in actors:
            myRow = [sheet.cell(row=i, column=1).value, actor.strip()]
            ws.append(myRow)
    except:
        pass

wb.save('D:\\Projects\\PythonProjects\\DataCleaning\\data\\Actors.xlsx')