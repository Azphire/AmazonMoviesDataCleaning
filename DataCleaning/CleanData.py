import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import re

source = load_workbook('D:\\Projects\\PythonProjects\\DataCleaning\\data\\Movies.xlsx')  # 读取xlsx中第一个sheet
sheet = source.get_sheet_by_name('Sheet')
rows = sheet.max_row
columns = sheet.max_column

# 存储输出文件
wb = Workbook()
ws = wb.active
ws.append(['asin', '电影名', '时长', '上映年份', '导演', '演员',
           '类别', '语言', '版本1', '版本2', '制片方', '用户评分'])
pattern1 = '\\(.*?\\)|\\[.*?]|(\|.*$)|(- Special Edition)'
pattern2 = ' - $'

# 遍历
# for i in range(13, 25):
for i in range(2, rows + 1):
    row = []
    for j in range(1, columns + 1):
        row.append(sheet.cell(row=i, column=j).value)

    # 正则表达式去括号
    title = re.sub(pattern1, "", str(row[1]))
    row[1] = re.sub(pattern2, "", title)
    row[1] = row[1].strip()

    # 时间简化为年份
    # if row[3]:
    #     date = row[3].split()
    #     row[3] = date[len(date) - 1]

    # 评分简化
    if row[11]:
        row[11] = row[11].split()[0]

    # 存入活动工作表
    ws.append(row)

wb.save('D:\\Projects\\PythonProjects\\DataCleaning\\data\\CleanedMovies.xlsx')
