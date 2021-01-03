from openpyxl import Workbook
from openpyxl import load_workbook
import sys  # 导入sys模块

sys.setrecursionlimit(3000)  # 将默认的递归深度修改为3000

mergeData = load_workbook(
    'D:\\Projects\\PythonProjects\\DataCleaning\\data\\mergedData.xlsx')  # 读取xlsx中第一个sheet
sheet = mergeData['Sheet1']
rows = sheet.max_row
columns = sheet.max_column

# 存储输出文件
wb = Workbook()
ws = wb.active
ws.append(['id', '电影名', '时长', '上映年份', '导演', '演员',
           '类别', '语言', '版本1', '版本2', '制片方', '用户评分', 'asin'])
movies = {}
for i in range(2, rows + 1):
    if sheet.cell(row=i, column=17).value == 0:
        continue
    myList = []
    asin = sheet.cell(row=i, column=2).value
    for j in range(3, 15):
        myList.append(sheet.cell(row=i, column=j).value)
    myList.append(sheet.cell(row=i, column=2).value)
    movies[asin] = myList
    if myList[11] == "master":
        continue
    else:
        movies[myList[11]][12] = movies[myList[11]][12] + "," + asin
amazonData = load_workbook(
    'D:\\Projects\\PythonProjects\\DataCleaning\\data\\Step1Result.xlsx')  # 读取xlsx中第一个sheet
amazonSheet = mergeData['Sheet1']
amazonRows = amazonSheet.max_row
amazonColumns = amazonSheet.max_column
# 整理第一次非主电影
movieIter = 0
# for i in range(amazonRows):
for i in range(2, amazonRows + 1):
    # for j in range(1, amazonColumns + 1):
    #     print(amazonSheet.cell(row=i, column=j).value)
    if amazonSheet.cell(row=i, column=14).value == "master":
        continue
    else:
        master = amazonSheet.cell(row=i, column=14).value  # 主电影asin
        try:
            asin = movies[master][11]  # 存储主电影的主电影列值
            if asin == "master":
                pass
            else:
                master = movies[asin][11]
            movies[master][12] = movies[master][12] + "," + amazonSheet.cell(row=i, column=2).value
        except:
            pass

movieId = 0
for movie in movies:
    if movies[movie][11] == "master":
        del movies[movie][11]
        movies[movie].insert(0, movieId)
        movieId += 1
        ws.append(movies[movie])

wb.save('D:\\Projects\\PythonProjects\\DataCleaning\\data\\FinalMovies.xlsx')
