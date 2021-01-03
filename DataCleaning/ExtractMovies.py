import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import re

source = load_workbook(
    'D:\\Projects\\PythonProjects\\DataCleaning\\data\\source.xlsx')  # 读取xlsx中第一个sheet
sheet = source.get_sheet_by_name('Sheet1')
rows = sheet.max_row
columns = sheet.max_column

# 存储输出文件
wb = Workbook()
ws = wb.active
ws.append(['asin', '电影名', '时长', '上映年份', '导演', '演员',
           '类别', '语言', '版本1', '版本2', '制片方', '用户评分'])
# patterns = ['Vol. ([0-9])+', 'Vol ([0-9])+', 'Volume ([0-9])+', 'TV Version', 'TV Show', 'Season [IV]']
# title中出现的能确定不是电影的正则表达式
patterns = '(Vol. ([0-9])+)|(Volume ([0-9])+)|(TV Version)|(TV Show)|(Season [IV])'

# 遍历
# for i in range(13, 25):
for i in range(2, rows + 1):

    row = []
    for j in range(1, columns + 1):
        row.append(sheet.cell(row=i, column=j).value)

    # print(row)

    director = row[4]
    # 通过是否有导演验证是否为电影
    if director is None:
        # print('无导演，不是电影')
        continue

    time = row[2]
    # 通过时长验证是否为电影
    if time is not None:
        # print('时间不为空')
        timeArray = time.split(' ')
        if len(timeArray) == 2:
            if timeArray[1] == 'minutes' or timeArray[1] == 'minute':
                minute = int(timeArray[0])
                # 时长不足半小时，不是电影
                if minute <= 30:
                    # print('时长不足，不是电影')
                    continue

    sort = row[6]
    # 通过类别筛选掉不是电影的
    if sort is not None:
        # print('类别不为空')
        if sort in ['TV', 'Sports', 'Exercise & Fitness', '3-6 Years', 'Instructional']:
            # print('类别不是电影')
            continue

    # 正则表达式拆分title排除非电影
    title = row[1]
    if title is not None:
        if re.search(patterns, title, re.I):
            # print('匹配到'+title)
            continue
        # match = False
        # for pattern in patterns:
        #     if re.search(pattern, title, re.I):
        #         # print('匹配到'+title)
        #         match = True
        #         continue
        # if match:
        #     continue

    # 是电影，存入活动工作表
    ws.append(row)

    # 是电影，存入movieData
    # movieData.append(row)


# for i in range(len(movieData)):
#     ws.append(movieData[i])
wb.save('D:\\Projects\\PythonProjects\\DataCleaning\\data\\Movies.xlsx')
