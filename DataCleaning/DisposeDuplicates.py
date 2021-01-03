
from openpyxl import Workbook
from openpyxl import load_workbook
from fuzzywuzzy import fuzz

source = load_workbook('D:\\Projects\\PythonProjects\\DataCleaning\\data\\CleanedMovies.xlsx')  # 读取xlsx中第一个sheet
sheet = source.get_sheet_by_name('Sheet')
rows = sheet.max_row
columns = sheet.max_column

# 存储输出文件
wb = Workbook()
ws = wb.active
ws.append(['asin', '电影名', '时长', '上映年份', '导演', '演员',
           '类别', '语言', '版本1', '版本2', '制片方', '用户评分', '主电影'])

# 遍历
# for i in range(2, 200):
for i in range(2, rows + 1):
    row = []
    for j in range(1, columns + 1):
        row.append(sheet.cell(row=i, column=j).value)
    # print(row)
    row.append('master')
    compareStart = i - 15  # 起始比较行
    if compareStart < 1:
        compareStart = 1
    master = True  # 默认是主电影
    for cmp in range(compareStart, i):
        # 判断title是否相等
        if str(sheet.cell(row=cmp, column=2).value).lower() == str(row[1]).lower():
            myProducers = row[4].split(', ')
            cmpProducers = sheet.cell(row=cmp, column=5).value.split(', ')
            # title相等，逐一判断导演
            for myProducer in myProducers:
                for cmpProducer in cmpProducers:
                    if fuzz.partial_ratio(myProducer, cmpProducer) > 85 or fuzz.partial_ratio(cmpProducer, myProducer) > 85:
                        # print(cmpProducer,myProducer)
                        master = False
                        row[12] = sheet.cell(row=cmp, column=1).value
                        break
                if not master:
                    break
            if not master:
                if row[5]:
                    myActors = row[5].split(', ')
                    cmpActors = ws.cell(row=cmp, column=6).value
                    for myActor in myActors:
                        # 若演员不在主电影中，添加进去
                        if fuzz.partial_ratio(cmpActors, myActor) < 98 and myActor:
                            add = str(cmpActors) + ', ' + str(myActor)
                            ws.cell(row=cmp, column=6, value=add)
                break

    # 存入活动工作表
    ws.append(row)

wb.save('D:\\Projects\\PythonProjects\\DataCleaning\\data\\Step1Result.xlsx')
